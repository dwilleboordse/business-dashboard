#!/usr/bin/env python3
"""
Convert the original "Business dashboard 2026 DDU.xlsx" into a JSON file that
matches the dashboard's AppState shape.

Usage:
    python3 scripts/excel_to_json.py [path-to-xlsx] [output.json]

Defaults:
    path-to-xlsx = ./Business dashboard 2026 DDU.xlsx
    output.json  = ./dashboard-import.json

Then in the running app, click "Import" in the header and select the JSON file.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

DUTCH_TO_EN = {
    "jan": "Jan", "feb": "Feb", "mrt": "Mar", "apr": "Apr",
    "mei": "May", "jun": "Jun", "jul": "Jul", "aug": "Aug",
    "sep": "Sep", "okt": "Oct", "nov": "Nov", "dec": "Dec",
}

# Map the Excel metric label (column B) to the dashboard metric id.
SOCIAL_METRIC_MAP = {
    "Followers Instagram":      "ig_followers",
    "Impressions Instagram":    "ig_impressions",
    "Posts Instagram":          "ig_posts",
    "Followers LinkedIn":       "li_followers",
    "Impressions LinkedIn":     "li_impressions",
    "Posts LinkedIn":           "li_posts",
    "Follower growth Twitter":  "tw_growth",
    "Impressions Twitter":      "tw_impressions",
    "Posts Twitter":            "tw_posts",
    "Subs YouTube":             "yt_subs",
    "Views YouTube":            "yt_views",
    "Posts YouTube":            "yt_posts",
    "Ad Spend":                 "ad_spend",
    "Call booked total":        "calls_total",
    "Call booked ads":          "calls_ads",
    "Gevoerde calls":           "calls_done",
    "Calls done":               "calls_done",
    "Proposals sent":           "proposals_sent",
    "Value proposals sent":     "proposals_value_sent",
    "Proposals accepted":       "proposals_accepted",
    "Value Proposals accepted": "proposals_value_accepted",
    "Reviews":                  "reviews",
    "Referrals":                "referrals",
    "Total clients":            "total_clients",
    "Clients left":             "clients_left",
    "Retention":                "retention",
}

ALL_METRIC_IDS = list(set(SOCIAL_METRIC_MAP.values())) + [
    "cpc_ads",       # computed
    "close_ratio",   # computed
]


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_socials(wb):
    ws = wb["Socials"]

    # Channel toggles live in rows 7-13, col B = bool, col C = name
    channels_default = {
        "Instagram": False, "Facebook": False, "YouTube": False,
        "TikTok": False, "Twitter": False, "LinkedIn": False, "Pinterest": False,
    }
    # Map case-insensitively from Excel ("Youtube", "Tiktok", "Linkedin") to canonical names
    canonical = {k.lower(): k for k in channels_default}
    for r in range(7, 14):
        flag = ws.cell(row=r, column=2).value
        name = ws.cell(row=r, column=3).value
        if name and isinstance(flag, bool):
            key = canonical.get(str(name).strip().lower())
            if key:
                channels_default[key] = flag

    # Weekly entries: rows 35-61. Col B=label, Cols C-T = Week 1..18, Col V = target
    weeks = [{mid: None for mid in ALL_METRIC_IDS} for _ in range(52)]
    targets = {}

    for r in range(35, 62):
        label = ws.cell(row=r, column=2).value
        if not label:
            continue
        metric_id = SOCIAL_METRIC_MAP.get(str(label).strip())
        if not metric_id:
            continue

        # Weeks 1..18 in cols C..T (3..20)
        for week_idx in range(18):
            col = 3 + week_idx
            v = num(ws.cell(row=r, column=col).value)
            if v is not None:
                weeks[week_idx][metric_id] = v

        target = num(ws.cell(row=r, column=22).value)
        if target is not None:
            targets[metric_id] = target

    return {"channels": channels_default, "weeks": weeks, "targets": targets}


def build_finance(wb):
    ws = wb["Omzet, kosten, winst"]

    monthly = {}
    # Rows 8..19 = jan..dec. B=label, C=goal, D=actual, H=costs
    for r in range(8, 20):
        label = ws.cell(row=r, column=2).value
        if not label:
            continue
        en_month = DUTCH_TO_EN.get(str(label).strip().lower())
        if not en_month:
            continue
        monthly[en_month] = {
            "goal": num(ws.cell(row=r, column=3).value) or 200000,
            "actual": num(ws.cell(row=r, column=4).value),
            "costs": num(ws.cell(row=r, column=8).value),
        }


    return {"monthly": monthly}


def build_emails(wb):
    ws = wb["Emails"]

    # Monthly subscribers: rows 7..18, B=month, C=subs
    monthly_subs = {m: None for m in DUTCH_TO_EN.values()}
    for r in range(7, 19):
        label = ws.cell(row=r, column=2).value
        if not label:
            continue
        en_month = DUTCH_TO_EN.get(str(label).strip().lower())
        if not en_month:
            continue
        monthly_subs[en_month] = num(ws.cell(row=r, column=3).value)

    # Per-email entries. Each month section starts at "jan"/"feb"/etc in column B.
    # Header row of the section has columns: B=month, C='Mailnaam', D='gem. open %', E='gem. click %', F='Unsubscribes' OR 'omzet', etc.
    # Rows below until next month-header are mail rows: B=date or 'Mail N' placeholder, C=name, D=open%, E=click%, F=unsubs/revenue
    emails_by_month = {m: [] for m in DUTCH_TO_EN.values()}

    current_month = None
    current_cols = None  # mapping of "open"/"click"/"unsubs"/"revenue" → column index

    for r in range(22, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        f = ws.cell(row=r, column=6).value
        g = ws.cell(row=r, column=7).value

        # Month header detection
        if isinstance(b, str) and b.strip().lower() in DUTCH_TO_EN and (
            (isinstance(c, str) and "mail" in c.lower()) or c is None
        ):
            current_month = DUTCH_TO_EN[b.strip().lower()]
            # Detect column meanings from headers d, e, f, g
            current_cols = {
                "open": 4 if isinstance(d, str) and "open" in d.lower() else None,
                "click": 5 if isinstance(e, str) and "click" in e.lower() else None,
                "unsubs": 6 if isinstance(f, str) and "unsub" in f.lower() else None,
                "revenue": 6 if isinstance(f, str) and ("omzet" in f.lower() or "revenue" in f.lower()) else None,
                "link": 7 if isinstance(g, str) and ("link" in g.lower()) else None,
            }
            continue

        if current_month is None:
            continue

        # Data row: c = email subject, d/e/f have stats, b is date or "Mail N" placeholder
        if not c or (isinstance(c, str) and c.strip().lower().startswith("mail ") and not d and not e and not f):
            # placeholder row, skip
            continue

        open_pct = num(ws.cell(row=r, column=current_cols["open"]).value) if current_cols.get("open") else None
        click_pct = num(ws.cell(row=r, column=current_cols["click"]).value) if current_cols.get("click") else None
        unsubs = num(ws.cell(row=r, column=current_cols["unsubs"]).value) if current_cols.get("unsubs") else None
        revenue = num(ws.cell(row=r, column=current_cols["revenue"]).value) if current_cols.get("revenue") else None
        link = ws.cell(row=r, column=current_cols["link"]).value if current_cols.get("link") else None

        # B may hold a real datetime, or a NL-style "day.month" decimal (e.g. 24.01 = Jan 24).
        date_str = ""
        if hasattr(b, "isoformat"):
            date_str = b.isoformat()[:10]
        elif isinstance(b, (int, float)):
            try:
                day_part, _, month_part = f"{b:.2f}".partition(".")
                day = int(day_part)
                month = int(month_part[:2])
                if 1 <= day <= 31 and 1 <= month <= 12:
                    date_str = f"2026-{month:02d}-{day:02d}"
            except (ValueError, IndexError):
                pass

        emails_by_month[current_month].append({
            "date": date_str,
            "name": str(c).strip(),
            "openPct": open_pct,
            "clickPct": click_pct,
            "unsubs": unsubs,
            "revenue": revenue,
            "link": str(link) if link else "",
        })

    return {"monthlySubscribers": monthly_subs, "emailsByMonth": emails_by_month}


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Business dashboard 2026 DDU.xlsx")
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("dashboard-import.json")

    if not xlsx.exists():
        print(f"Could not find {xlsx}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(xlsx), data_only=True)

    state = {
        "year": 2026,
        "socials": build_socials(wb),
        "finance": build_finance(wb),
        "emails": build_emails(wb),
    }

    out.write_text(json.dumps(state, indent=2, ensure_ascii=False))
    print(f"Wrote {out}")
    print(f"  Socials weeks with data: {sum(1 for w in state['socials']['weeks'] if any(v is not None for v in w.values()))}")
    print(f"  Finance months with data: {sum(1 for m in state['finance']['monthly'].values() if m['actual'] is not None)}")
    print(f"  Email entries: {sum(len(es) for es in state['emails']['emailsByMonth'].values())}")


if __name__ == "__main__":
    main()
